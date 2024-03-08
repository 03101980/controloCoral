"""Ce fichier contient la page principal qui charge toute l'aplication.
Il s'appuie sur les fichiers function qui contient toutes les fonctions de l'application:
"""
import sys, os
from  PyQt6 import uic, QtGui
from PyQt6.QtWidgets import *
from PyQt6.QtGui import * 
from PyQt6.QtCore import *
from function import *
from openpyxl import load_workbook

basedir = os.path.dirname(__file__)

try:
    from ctypes import windll  # Only exists on Windows.
    myappid = 'maestro.com'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except ImportError:
    pass

class Coro_ccda(QMainWindow):
    """essa é a base do applicativo, contem todo o funcionamento do aplicativo
    as ligações dos butons e alguns funções"""
    def __init__(self):    
        super(Coro_ccda, self).__init__()
        uic.loadUi(r'PaginaPrincipal.ui',self)
        icon = QtGui.QIcon(os.path.join(basedir, 'imagens/logo.ico'))
        self.setWindowIcon(icon)
        #eliminar bara de titulo 
        # self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        # self.setWindowOpacity(1)
        # self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        #sizegrip
        # self.gripSize = 10
        # self.grip = QtWidgets.QSizeGrip(self)
        # self.grip.resize(self.gripSize, self.gripSize)

        # self.click_position = QPoint()
        # self.frame_titre.mouseMoveEvent = self.mover
        #colocar sombra
        self.sombra(self.bt_actualizar)
        self.sombra(self.bt_pag_membros)
        self.sombra(self.bt_pag_cotas)
        self.sombra(self.bt_pag_caixa)
        self.sombra(self.bt_pag_actualiza)
        self.sombra(self.bt_eliminar)
        self.sombra(self.bt_adicionar)
        self.sombra(self.bt_editar)
        self.sombra(self.LcdData)
        ########## as ligações dos butões##################
        ###################################################
        self.bt_menu1.clicked.connect(self.moverMenu)
        self.Bt_menu2.clicked.connect(self.moverMenu)
        self.bt_pag_membros.clicked.connect(self.pagina_membros)
        self.bt_pag_caixa.clicked.connect(self.pagina_caixa)
        self.bt_pag_cotas.clicked.connect(self.pagina_cota)
        self.bt_pag_actualiza.clicked.connect(self.actualizarDadoGrupo)
        self.bt_adicionar.clicked.connect(self.moverFormulario)
        self.bt_cancelar.clicked.connect(self.fechar)
        self.bt_validar.clicked.connect(self.validar_registo)
        self.bt_editarDetalhe.clicked.connect(self.editarDadosMembro)
        self.bt_novoPagamento.clicked.connect(self.formularioPagamento)
        self.Bt_carregaFoto.clicked.connect(self.caregar_foto)
        self.bt_actualizar.clicked.connect(self.carregar_dados)
        self.bt_eliminar.clicked.connect(self.desactivarMembros)
        self.bt_eliminar_reciclage.clicked.connect(self.eliminarMembro)
        self.bt_restaurar.clicked.connect(self.activar_membro)
        self.bt_eliminarDetalha.clicked.connect(self.eliminarMembro)
        self.bt_restaurarDetalha.clicked.connect(self.activar_membro)
        self.Bt_detalho.clicked.connect(self.detalho)
        self.bt_avancar.clicked.connect(self.actualizarListaCota)
        self.bt_recuar.clicked.connect(self.actualizarListaCota)
        self.bt_editar.clicked.connect(self.editarDadosMembro)
        self.Bt_saida.clicked.connect(self.formularioDoMovimento)
        self.Bt_entrada.clicked.connect(self.formularioDoMovimento)
        self.bt_criarficheiroExcelCaixa.clicked.connect(self.criar_excel_caixa)
        self.Bt_criarFicheiroExcelMembros.clicked.connect(self.criar_excel_membros)
        self.Bt_criarFicheiroExcelCota.clicked.connect(self.criar_excel_cota)
        ########## LineEdit ou caixas de texto ########################
        # para se connectar as funções enquanto digitalizar ###########
        self.LinEdit_filtrar.textChanged.connect(self.carregar_dados)
        self.LinEd_pesquisa.textChanged.connect(lambda: self.visualizar_cotas(0))
        ################# ComboBox ################ coneção as funções ##############
        self.comboBox_SelectEstado.currentIndexChanged.connect(self.carregar_dados)
        self.ComBox_membrosActivadoCota.currentIndexChanged.connect(lambda: self.visualizar_cotas(0))
        self.comboBox_filtroCaixa.currentIndexChanged.connect(self.visualizarCaixa)
        ################# DateEdit ################ coneção as funções ##############
        self.dateEdit_Data_inicial.dateChanged.connect(self.visualizarCaixa)
        self.dateEdit_Data_fim.dateChanged.connect(self.visualizarCaixa)
        ############## esconder alguns butões###############
        self.Bt_menu2.hide() #esconder o butão do menu
        self.LabelAno.hide()
        self.bt_restaurar.hide()
        self.bt_eliminar_reciclage.hide()
        self.Label_valorCota.hide()
        self.bt_recuar.setEnabled(False)
        #inicialização de table dos dados dos membros e ds cotas
        self.table_membros.setColumnCount(6)  # Número de colunas
        self.table_membros.setHorizontalHeaderLabels(["Nome", "Paróquia", "Morada", "Contacto", "Voz","id",])
        self.table_cotas.setColumnCount(5)  # Número de colunas
        self.table_cotas.setHorizontalHeaderLabels(["Nome","Setembro", "Outubro", "Novembro","Dezembro"])
        self.table_membros.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        for num in range(1,6):
            self.table_membros.horizontalHeader().resizeSection(num, 100)

        ###### inicializar funções para carregar dados do membros e das cotas########
        self.carregar_dados()
        self.visualizar_cotas(1)
        self.demaragem_loop()
        self.lineEdit()

        self.contar = 0 # contage para a fução actualizarLista Cota afim de mudar o cabeçalho de mese
        ### Inicializar os variaveis dos mese para a pesquiça no banco de dados
        self.mes_pagamento_inf = '8'
        self.mes_pagamento_sup = '13'
        self.action = 0 # para distenguer l'edition de dados et l'adition de novo membros 
        self.id_membros = 0    
        #self.esconderFormulario()
    #     self.DatEdit_date_ingresso.dateChanged.connect(self.update_date)
    #     self.update_date(self.DatEdit_date_ingresso.date())
    # def update_date(self, da):
    #     pass
    #     #A função update_date será chamada sempre que a data for alterada
    #     print(f'Data Selecionada: {da.toString("yyyy-MM-dd")}')
    #     print(f'Data Selecionada: {da.toString("dd-MM-yyyy")}')
    #     print(f'Data Selecionada: {self.DatEdit_date_ingresso.date().toString("yyyy-MM-dd")}')
    def lineEdit(self):
        bairro = [
            'Apple',
            'Apricot',
            'Banana',
            'Rose apple',
            'Starfruit',
            'Strawberries',
            'Water apple',
        ]
        # Criar um QCompleter com as sugestões
        completer = QCompleter(bairro, self.LineEdir_morada)
        # Definir o QCompleter no QLineEdit
        self.LineEdir_morada.setCompleter(completer)
        self.LineEdit_contacto.setInputMask('+244_999_999_999')
    def demaragem_loop(self):
        # Use um loop while para repetir até que a função demaragem seja bem-sucedida
        # Retorna True se a demaragem for bem-sucedida, False caso contrário
        dados = caregarDadoGrupo()
        if dados:
            for dado in dados:
                self.Label_titreDoGroupe.setText(str(dado[1]))
                self.Label_valorCota.setText(str(dado[3]))
        else:
            # Se a demaragem não for bem-sucedida, abra o formulário do grupo novamente
            self.formularioDogrupoo()

    def criar_excel_cota(self):
        # Abrir um arquivo Excel existente, edita-ló com a lista das cotas e salvar de novo
        file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            try:
                wb = load_workbook(file_name)
                # Conectar ao banco de dados SQLite e carregar ainda os nomes
                arquivo = carregar_cotas_nome(self.LinEd_pesquisa.text(), self.ComBox_membrosActivadoCota.currentText()) 
                ws = wb.active
                #Limpar dados existentes na planilha
                for row in ws.iter_rows(min_row=6, min_col=2): #assim, com (5) a limpeza começa na sexta linha affi de deixar o cabeçalho permanecer
                    for cell in row:
                        cell.value = None
                #preenchimento
                for row_num, row_data in enumerate(arquivo, start=1):
                    row_num = row_num + 5 # o preenchemento começa na linha 6
                    ws.cell(row=row_num, column=2, value=row_data[1])
                    valor = carregar_cotas_valor(row_data[0], int(self.mes_pagamento_inf), int(self.mes_pagamento_sup), int(self.LabelAno.text()))
                    for col_num, col_value in enumerate(valor, start=1):
                        num = 0               
                        ws.cell(row=row_num, column=col_num + 2, value=col_value[num])
                        num += 1
                alerta = QMessageBox.question(self, 'Alerta', f'Deseja salvar o arquivo {file_name} ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
                if alerta == QMessageBox.StandardButton.Yes:
                    # Salvar as alterações no arquivo Excel
                     # Obter um nome de arquivo e localização personalizados antes de salvar
                    save_file_name, _ = QFileDialog.getSaveFileName(self, "Salvar Excel", "", "Excel Files (*.xlsx);;All Files (*)")
                    if save_file_name:
                        wb.save(save_file_name)
                        self.exibir_mensagem('Excel', f'Arquivo salvo neste diretório: {save_file_name}')
                    else:
                        pass
                else:
                    pass
            except Exception as e:
                self.exibir_mensagem('Excel',f'{e} erro')
    def criar_excel_membros(self):
        # Abrir um arquivo Excel existente
        #file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            try:
                wb = load_workbook(file_name)
                # Conectar ao banco de dados SQLite
                arquivo = carregar_dados_excel()
                # Preencher o arquivo Excel com dados do banco de dados
                ws = wb.active
                for row in ws.iter_rows(min_row=6, min_col=2):
                    for cell in row:
                        cell.value = None

                for row_num, row_data in enumerate(arquivo, start=1):
                    row_num = row_num + 5
                    for col_num, col_value in enumerate(row_data, start=1):                       
                        ws.cell(row=row_num, column=col_num + 1, value=col_value)
                alerta = QMessageBox.question(self, 'Alerta', f'Deseja salvar as alterações feita no arquivo {file_name} ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
                if alerta == QMessageBox.StandardButton.Yes:
                    # Salvar as alterações no arquivo Excel
                    save_file_name, _ = QFileDialog.getSaveFileName(self, "Salvar Excel", "", "Excel Files (*.xlsx);;All Files (*)")
                    if save_file_name:
                        wb.save(save_file_name)
                        self.exibir_mensagem('Excel', f'Arquivo salvo neste diretório: {save_file_name}')
                    else:
                        pass
                else:
                    pass
            except Exception as e:
                self.exibir_mensagem('Excel',f'{e} erro')

    def criar_excel_caixa(self):
        # Abrir um arquivo Excel existente
        file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            try:
                wb = load_workbook(file_name)
                # Conectar ao banco de dados SQLite
                arquivo = carregar_dadosCaixaExcel(self.comboBox_filtroCaixa.currentText(), 
                                        self.dateEdit_Data_inicial.date().toString("yyyy-MM-dd"), 
                                        self.dateEdit_Data_fim.date().toString("yyyy-MM-dd"))
                # Preencher o arquivo Excel com dados do banco de dados
                ws = wb.active
                # Limpar todos os valores nas células
                for row in ws.iter_rows(min_row=6, min_col=2):
                    for cell in row:
                        cell.value = None
                total_cota = 0
                linha = 6 # determinar a linha 5 como o começo
                for row_num, row_data in enumerate(arquivo, start=1):
                    if row_data[1] == 'Pagamento Cota':
                        total_cota += row_data[2]
                    else:
                        for col_num, col_value in enumerate(row_data, start=1):
                            #utilizamos o variavel linha emvés de row_num para evitar que pula a linha 
                            # no momento de adicionar os pagamento de cotas no if em cima                                             
                            ws.cell(row=linha, column=col_num + 1, value=col_value)
                        linha += 1
                # em fim, preenchemos o total do pagamento
                ws.cell(row=linha, column= 2, value='-----  Total  -----')
                ws.cell(row=linha, column= 3, value='Pagamento Cota')
                ws.cell(row=linha, column= 4, value=total_cota)
                alerta = QMessageBox.question(self, 'Alerta', f'Deseja salvar as alterações feitas no arquivo: {file_name} ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
                if alerta == QMessageBox.StandardButton.Yes:
                    # Salvar as alterações no arquivo Excel
                    save_file_name, _ = QFileDialog.getSaveFileName(self, "Salvar Excel", "", "Excel Files (*.xlsx);;All Files (*)")
                    if save_file_name:
                        wb.save(save_file_name)
                        self.exibir_mensagem('Excel', f'Arquivo salvo neste diretório: {save_file_name}')
                    else:
                        pass
                else:
                    pass
            except Exception as e:
                self.exibir_mensagem('Excel',f'{e} erro')

    def eliminarMembro(self):
        selecao = self.table_membros.selectedItems()
        if selecao:
            dados = []  # Inicialize a lista fora do loop
            for item in selecao:
                # Adicione o texto de cada item à lista
                dados.append(item.text())
            alerta = QMessageBox.question(self, 'Alerta', f'Deseja realmente eliminar {dados[0]} da lista dos membros ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if alerta == QMessageBox.StandardButton.Yes:       
                resposta = eliminarMembros(dados[5])
                if resposta == 1:
                    self.exibir_mensagem('Banco de dados',f'{dados[0]} foi eliminado com sucesso')
                else:
                    self.exibir_mensagem('Banco de dados',f'Erro: {resposta}, para eliminar {dados[0]}')
            else:
                pass  # Ignorar o evento de fechamento
        else:
            self.exibir_mensagem('Edição','Nenhum dado seleccionado')
            pass
        self.carregar_dados()

    def editarDadosMembro(self):
        self.action = 1 # un signifie l'edition et 0 c'est adition de novo membro
        self.stackedWidget.setCurrentWidget(self.pag_membros) # cao etiver na pagina de detalho
        selecao = self.table_membros.selectedItems()
        if selecao:
            dados = []  # Inicialize a lista fora do loop
            for item in selecao:
                # Adicione o texto de cada item à lista
                dados.append(item.text())
                # Se quiser apenas o terceiro elemento, use dados[2]
            detalho = carregar_todos_dados(dados[5])
            for resultat in detalho:
                self.id_membros = resultat[0]
                self.LineEdit_nome.setText(str(resultat[1]))
                self.ComBox_paroquia.setCurrentText(str(resultat[3]))
                self.LineEdir_morada.setText(str(resultat[4]))
                self.LineEdit_contacto.setText(str(resultat[5]))
                self.ComboBox_Voz.setCurrentText(str(resultat[6]))
                self.ComboBox_estado.setCurrentText(str(resultat[9]))
                if resultat[2] == "F":
                    self.Sexo_fem.setChecked(True)
                if len(str(resultat[7])) < 5:
                    self.label_imagem.setText('Nenhuma foto encotrada')
                else:
                    foto = QPixmap()
                    foto.loadFromData(resultat[7], "PNG")
                    self.label_imagem.setPixmap(foto)         
            self.moverFormulario()
        else:
            self.exibir_mensagem('Edição','Nenhum dado seleccionado')
            pass       

    def visualizarCaixa(self):
        self.table_caixa.setRowCount(0)
        self.table_caixa.setColumnCount(5)  # Número de colunas
        self.table_caixa.setHorizontalHeaderLabels(["Data","Motivo", "Entrada", "Saída","Saldo"])
        self.table_caixa.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_caixa.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        for num in range(2,5):
            self.table_caixa.horizontalHeader().resizeSection(num, 100)

        resultado = carregar_dadosCaixa(self.comboBox_filtroCaixa.currentText(), 
                                        self.dateEdit_Data_inicial.date().toString("yyyy-MM-dd"), 
                                        self.dateEdit_Data_fim.date().toString("yyyy-MM-dd"))
        total_entrada = 0
        total_saida = 0
        for row_num, row_data in enumerate(resultado):
            self.table_caixa.insertRow(row_num)
            total_entrada += row_data[2]
            total_saida += row_data[3]
            for linha_num, linha_data in enumerate(row_data):          
                item = QTableWidgetItem(str(linha_data))
                self.table_caixa.setItem(row_num, linha_num, item)
                if row_num % 2 == 0:
                    item.setBackground(QColor(167, 117, 0))  # Cor de fundo para linhas pares
                else:
                    item.setBackground(QColor(255, 205, 0))  # Cor de fundo para linhas ímpares
        if not resultado:
            pass            
        else:
            self.table_caixa.insertRow(row_num + 1) # aumentar uma linha na tabela
            # carregar o dado
            total = QTableWidgetItem('Total')
            entrada = QTableWidgetItem(str(total_entrada))
            saida = QTableWidgetItem(str(total_saida))     
            saldo = QTableWidgetItem(str(total_entrada - total_saida))
            # inerir dado na tabela                         
            self.table_caixa.setItem(row_num + 1, 1, total)
            self.table_caixa.setItem(row_num + 1, 2, entrada)
            self.table_caixa.setItem(row_num + 1, 3, saida)
            self.table_caixa.setItem(row_num + 1, 4, saldo)
            # Core da linha
            total.setBackground(QColor(170, 0, 0))  
            entrada.setBackground(QColor(170, 0, 0))  
            saida.setBackground(QColor(170, 0, 0))  
            saldo.setBackground(QColor(167, 0, 0))

    def actualizarListaCota(self):
        buton = self.sender()
        table_um = ["Nome","Setembro", "Outubro", "Novembro","Dezembro"]
        table_dois = ["Nome", "Janeiro", "Fevereiro", "Março", "Abril", "Maio","Junho"]
        table_tres = ["Nome", "Julho", "Agosto", "Setembro", "Outubro", "Novembro","Dezembro"]
        ano = int(self.LabelAno.text())
        if isinstance(buton, QPushButton):
            if buton.text() == "A seguir":
                self.contar +=  1
                self.table_cotas.setColumnCount(6)  # Número de colunas
                if self.contar % 2 == 1:               
                    self.table_cotas.setHorizontalHeaderLabels(table_dois)
                    self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                    for num in range(1,6):
                        self.table_cotas.horizontalHeader().resizeSection(num, 100)
                    self.mes_pagamento_inf = '0'
                    self.mes_pagamento_sup = '7'
                    ano += 1            
                else:
                    self.table_cotas.setHorizontalHeaderLabels(table_tres)
                    self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                    for num in range(1,6):
                        self.table_cotas.horizontalHeader().resizeSection(num, 100)
                    self.mes_pagamento_inf = '6'
                    self.mes_pagamento_sup = '13'
                    ano = ano 
                self.bt_recuar.setEnabled(True)
            elif buton.text() == "recuar":
                self.contar -= 1
                if self.contar > 0:
                    if self.contar % 2 == 0:               
                        self.table_cotas.setHorizontalHeaderLabels(table_tres)
                        self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                        for num in range(1,6):
                            self.table_cotas.horizontalHeader().resizeSection(num, 100)
                        self.mes_pagamento_inf = '6'
                        self.mes_pagamento_sup = '13'  
                        ano -= 1                                 
                    else:
                        self.table_cotas.setHorizontalHeaderLabels(table_dois)
                        self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                        for num in range(1,6):
                            self.table_cotas.horizontalHeader().resizeSection(num, 100)
                        self.mes_pagamento_inf = '0'
                        self.mes_pagamento_sup = '7'
                else:
                    self.table_cotas.setColumnCount(5)  # Número de colunas
                    self.table_cotas.setHorizontalHeaderLabels(table_um)
                    self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                    for num in range(1,5):
                        self.table_cotas.horizontalHeader().resizeSection(num, 100)
                    self.mes_pagamento_inf = '8'
                    self.mes_pagamento_sup = '13'
                    self.bt_recuar.setEnabled(False)
                    ano = 2023
            self.LabelAno.setText(str(ano))
            self.LcdData.display(self.LabelAno.text())
            self.visualizar_cotas(0)
        else:
            self.table_cotas.setColumnCount(5)  # Número de colunas
            self.table_cotas.setHorizontalHeaderLabels(table_um)
            self.mes_pagamento_inf = '8'
            self.mes_pagamento_sup = '13'
        
    def visualizar_cotas(self, indice):
        cotas = carregar_cotas_nome(self.LinEd_pesquisa.text(), self.ComBox_membrosActivadoCota.currentText())
        # Limpar a tabela antes de preenchê-la
        self.table_cotas.setRowCount(0)
        if indice == 1:
            self.table_cotas.setColumnCount(5)  # Número de colunas
            self.table_cotas.setHorizontalHeaderLabels(["Nome","Setembro", "Outubro", "Novembro","Dezembro"])
            # Exemplo de redimensionamento automático de colunas com base no conteúdo
            self.table_cotas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            for num in range(1,5):
                self.table_cotas.horizontalHeader().resizeSection(num, 100)
            self.mes_pagamento_inf = '8'
            self.mes_pagamento_sup = '13'
        # Adicionar linhas à tabela
        for row_num, row_data in enumerate(cotas):
            self.table_cotas.insertRow(row_num)
            item = QTableWidgetItem(str(row_data[1]))
            self.table_cotas.setItem(row_num, 0, item) # inserir os nomes
            if row_num % 2 == 0:
                item.setBackground(QColor(167, 117, 0))  # Cor de fundo para linhas pares
            else:
                item.setBackground(QColor(255, 205, 0))  # Cor de fundo para linhas ímpares
            valor = carregar_cotas_valor(row_data[0], int(self.mes_pagamento_inf), int(self.mes_pagamento_sup), int(self.LabelAno.text()))
            col_num = 1
            for col_data in valor:            
                for valorPago in col_data:
                    if not valorPago:
                        pass
                    else:
                        item = QTableWidgetItem(str(valorPago))
                        self.table_cotas.setItem(row_num, col_num, item)
                        if row_num % 2 == 0:
                            item.setBackground(QColor(167, 117, 0))  # Cor de fundo para linhas pares
                        else:
                            item.setBackground(QColor(255, 205, 0))  # Cor de fundo para linhas ímpares
                        col_num += 1
                    
    def carregar_dados(self):
        if self.comboBox_SelectEstado.currentText() == "Reciclagem":
            self.bt_restaurar.show()
            self.bt_eliminar_reciclage.show()
        else:
            self.bt_restaurar.hide()
            self.bt_eliminar_reciclage.hide()
        # Carregar dados do banco de dados
        dados_membros = carregar_dados_do_banco(self.LinEdit_filtrar.text(), self.comboBox_SelectEstado.currentText())
        # Preencher o QTableWidget com os dados
        self.preencher_tabela(dados_membros)
        
    def preencher_tabela(self, dados):
        # Limpar a tabela antes de preenchê-la
        self.table_membros.setRowCount(0)
        # Adicionar linhas à tabela
        for row_num, row_data in enumerate(dados):
            self.table_membros.insertRow(row_num)
            for col_num, col_data in enumerate(row_data):
                item = QTableWidgetItem(str(col_data))
                self.table_membros.setItem(row_num, col_num, item)
                 # Definir a cor de fundo para linhas pares e ímpares
                if row_num % 2 == 0:
                    item.setBackground(QColor(167, 117, 0))  # Cor de fundo para linhas pares
                else:
                    item.setBackground(QColor(255, 205, 0))  # Cor de fundo para linhas ímpares

        # Definir a cor de fundo para o cabeçalho
        header = self.table_membros.horizontalHeader()
        for col_num in range(self.table_membros.columnCount()):
            cabecalho_item = self.table_membros.horizontalHeaderItem(col_num)
            cabecalho_item.setBackground(QColor(150, 150, 150))  # Cor de fundo para o cabeçalho

    # def minimizar(self):
    #     self.showMinimized()
    # def maximizar(self):
    #     self.showMaximized()
    #     self.bt_normal.show()
    #     self.bt_max.hide()
    # def normalizar(self):
        # self.showNormal()
        # self.bt_normal.hide()
        # self.bt_max.show()
    def caregar_foto(self):
        foto = QFileDialog.getOpenFileName(
            filter="Image JPG(*.jpg);;Image PNG(*.png)")[0]
        if foto:
            pixmap_image = QPixmap(foto).scaled(300, 400, aspectRatioMode=Qt.AspectRatioMode.KeepAspectRatio, transformMode=Qt.TransformationMode.SmoothTransformation)
            self.label_imagem.setPixmap(pixmap_image)
            
    # def moverpagina(self, event):
    #     rect = self.rect()
    #     self.grip.move(rect.right() - self.gripSize, rect.bottom() - self.gripSize)
    
    # def mousePressEvent(self, event):
    #     if event.button() == Qt.MouseButton.LeftButton:
    #         self.click_position = event.globalPosition()

    # def mover(self, event):
    #     if self.isMaximized() == False:
    #         if event.buttons() == Qt.MouseButton.LeftButton:
    #             new_pos = QPointF(self.pos()) + event.globalPosition() - self.click_position
    #             self.move(int(new_pos.x()), int(new_pos.y()))
    #             self.click_position = event.globalPosition()
    #             event.accept()
    #     if event.globalPosition().y() <= self.gripSize:
    #         self.showMaximized()
    #         self.bt_normal.hide()
    #         self.bt_max.show()
    #     else:
    #         self.showNormal()
    #         self.bt_normal.show()
    #         self.bt_max.hide()

    def sombra(self, frame):
        sombra = QGraphicsDropShadowEffect(self)
        sombra.setBlurRadius(20)
        sombra.setXOffset(1)
        sombra.setYOffset(1)
        sombra.setColor(QColor(255,180,0,255))
        frame.setGraphicsEffect(sombra)

    def moverMenu(self):
        if True:
            width = self.frame_menu.width()
            normal = 0
            if  width == 0:
                extender = 170
                #mudança de butão
                self.bt_menu1.show()
                self.Bt_menu2.hide()
            else:
                #mudança de butão
                extender = normal
                self.bt_menu1.hide()
                self.Bt_menu2.show()
                #reglage e demaragen de animação
            self.animation = QPropertyAnimation(self.frame_menu, b"maximumWidth")
            self.animation.setStartValue(width)
            self.animation.setEndValue(extender)
            self.animation.setDuration(2000)
            self.animation.start()
                        
    def moverFormulario(self):
        if True:
            height = self.frameFormulario.height()
            normal = 0
            if  height == 0:
                extender = 500
                self.bt_adicionar.setEnabled(False)
                self.bt_actualizar.setEnabled(False)
                self.bt_eliminar.setEnabled(False)
                self.bt_editar.setEnabled(False)
            else:
                extender = normal
                self.bt_adicionar.setEnabled(True)
                self.bt_actualizar.setEnabled(True)
                self.bt_eliminar.setEnabled(True)
                self.bt_editar.setEnabled(True)
            self.animation = QPropertyAnimation(self.frameFormulario, b"maximumHeight")
            self.animation.setStartValue(height)
            self.animation.setEndValue(extender)
            self.animation.setEasingCurve(QEasingCurve.Type.OutBounce)
            self.animation.setDuration(2000)
            self.animation.start()
            # self.anim_2 = QPropertyAnimation(self.frameFormulario, b"opacity")
            # self.anim_2.setStartValue(10)
            # self.anim_2.setEndValue(0)
            # self.anim_2.setDuration(2500)
            # self.anim_group = QParallelAnimationGroup()
            # self.anim_group.addAnimation(self.animation)
            # self.anim_group.addAnimation(self.anim_2)
            # self.anim_group.start()
    def desactivarMembros(self):
        selecao = self.table_membros.selectedItems()
        if selecao:
            dados = []  # Inicialize a lista fora do loop
            for item in selecao:
                # Adicione o texto de cada item à lista
                dados.append(item.text())
                # Se quiser apenas o terceiro elemento, use dados[2]
            self.id_membros = dados[5]
            alerta = QMessageBox.question(self, 'Alerta', f'Deseja realmente enviar {dados[0]} para lixeira ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if alerta == QMessageBox.StandardButton.Yes:
                resultado = activation_membro(dados[5], 'desactivado')                            
                if resultado == 1:
                    self.exibir_mensagem("Registo de Membros", "Membro desactivado com sucesso!")           
                else:
                    self.exibir_mensagem("Registo de Membros", "Erro ao desactivado membro: {}".format(resultado))
        else:
            self.exibir_mensagem("Registo de Membros", "Nenhum dado selecionado!")
        self.carregar_dados()
    def activar_membro(self):
        selecao = self.table_membros.selectedItems()
        if selecao:
            dados = []  # Inicialize a lista fora do loop
            for item in selecao:
                # Adicione o texto de cada item à lista
                dados.append(item.text())
                # Se quiser apenas o terceiro elemento, use dados[2]
            self.id_membros = dados[5]
            alerta = QMessageBox.question(self, 'Alerta', f'Deseja realmente restaurar {dados[0]} ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if alerta == QMessageBox.StandardButton.Yes:
                resultado = activation_membro(dados[5], 'Activado')                            
                if resultado == 1:
                    self.exibir_mensagem("Registo de Membros", "Membro activado com sucesso!")           
                else:
                    self.exibir_mensagem("Registo de Membros", "Erro ao activado membro: {}".format(resultado))
        else:
            self.exibir_mensagem("Registo de Membros", "Nenhum dado selecionado!") 
        self.carregar_dados()
    def detalho(self):       
        selecao = self.table_membros.selectedItems()
        if selecao:
            dados = []  # Inicialize a lista fora do loop
            for item in selecao:
                # Adicione o texto de cada item à lista
                dados.append(item.text())
                # Se quiser apenas o terceiro elemento, use dados[2]
            self.id_membros = dados[5]
            self.action = 1
            detalho = carregar_todos_dados(dados[5])
            for resultat in detalho:
                self.LinEd_nome.setText(str(resultat[1]))
                self.LinEd_sexo.setText(str(resultat[2]))
                self.LinEd_paroquia.setText(str(resultat[3]))
                self.LinEd_morada.setText(str(resultat[4]))
                self.LinEd_contacto.setText(str(resultat[5]))
                self.LinEd_voz.setText(str(resultat[6]))
                self.LinEd_dataIngresso.setText(str(resultat[8]))

                if len(str(resultat[7])) > 3:
                    foto = QPixmap()
                    foto.loadFromData(resultat[7], "PNG")
                    self.Label_foto.setPixmap(foto)
                else:
                    self.Label_foto.setText('Nenhuma foto encotrada')
            self.stackedWidget.setCurrentWidget(self.page_detalho)
        else:
            self.exibir_mensagem('Detalho de dados','Nenhum item selecionado.')
               
    def esconderFormulario(self):
        self.animation = QPropertyAnimation(self.frameFormulario, b"maximumHeight")
        self.animation.setEndValue(0)
        self.animation.setDuration(50)
        self.animation.start()
        
    def pagina_caixa(self):
        self.stackedWidget.setCurrentWidget(self.pag_caixa)
        self.visualizarCaixa()
    def pagina_membros(self):
        self.stackedWidget.setCurrentWidget(self.pag_membros)
        #self.animation_pagina()
        self.carregar_dados()
    def pagina_cota(self):
        self.stackedWidget.setCurrentWidget(self.pag_cotas)
        #self.animation_pagina()
        self.visualizar_cotas(1)

    def fechar(self):
        #if isinstance(self.sender(), QPushButton): # aqui on verifie se le signal est de QPushButton.
        buton = self.sender()
        if buton.text() == "Cancelar":
            reply = QMessageBox.question(self, 'Titulo', 'Deseja realmente fechar e abandonar esses dados ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:       
                self.limpar_formulario()  # Aceitar o evento de fechamento
                self.moverFormulario()
                self.carregar_dados()
            else:
                pass  # Ignorar o evento de fechamento
        else:
            reply = QMessageBox.question(self, 'Titulo', 'Deseja realmente fechar e abandonar o aplicativo ?',
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)

            if reply == QMessageBox.StandardButton.Yes:
                    
                self.close()  # Aceitar o evento de fechamento
            else:
                pass  # Ignorar o evento de fechamento
    def limpar_formulario(self):
        self.LineEdit_nome.clear()
        #self.ComBox_paroquia.setIndex(0)
        self.LineEdir_morada.clear()
        self.LineEdit_contacto.clear()
        #self.ComboBox_Voz.setIndex(0)
    def exibir_mensagem(self, titulo, mensagem):
        msg_box = QMessageBox()
        msg_box.setWindowTitle(titulo)
        msg_box.setText(mensagem)
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.exec()

    def validar_registo(self):
        nome = self.LineEdit_nome.text()
        paroquia = self.ComBox_paroquia.currentText()
        morada = self.LineEdir_morada.text()
        contacto = self.LineEdit_contacto.text()
        voz = self.ComboBox_Voz.currentText()
        foto = self.label_imagem.pixmap()
        data = self.DatEdit_date_ingresso.date().toString("yyyy-MM-dd")
        if self.Sexo_fem.isChecked():
            sexo = "F"
        else:
            sexo = "M"
        #convrtit foto Pixmap en bytes
        if foto:
            bArray = QByteArray()
            buff = QBuffer(bArray)
            buff.open(QIODevice.OpenModeFlag.WriteOnly)
            foto.save(buff, "PNG")
            buff.close()
        resultado = inserir_membro(nome, sexo, paroquia, morada, contacto, voz, bArray, self.action, self.id_membros, data, 'Activado')
        if resultado == 1:
            self.exibir_mensagem("Registo de Membros", "Membro inserido com sucesso!")                    
        else:
            self.exibir_mensagem("Registo de Membros", "Erro ao inserir membro: {}".format(resultado))
        self.action = 0
        self.limpar_formulario() 
        self.moverFormulario()
        self.carregar_dados()

    def formularioPagamento(self):
        dialog = formulario_cotas(self)
        result = dialog.exec()
        if result == QDialog.DialogCode.Accepted:
            if dialog.linEd_valor.text() != "":
                valor = dialog.linEd_valor.text()
                index = dialog.CombBox_nome.currentIndex()
                valor_cota = self.Label_valorCota.text()
                if index != -1:
                    id_membro = dialog.CombBox_nome.itemData(index, Qt.ItemDataRole.UserRole)
                caixa, pagamento = validar_cota(id_membro, valor, valor_cota)
                if pagamento == 1:
                    self.exibir_mensagem("Pagamento Cota", "Pagamento efectuado com sucesso!")
                    if caixa == 1:
                        self.exibir_mensagem("Caixa", "Valor guardado com sucesso!")
                    else:
                        self.exibir_mensagem("Caixa", f"Erro para quardar!{caixa}")            
                    self.visualizar_cotas(1)
                else:
                    self.exibir_mensagem("Pagamento Cota", f"Erro do pagamento: {pagamento}")
                    if caixa == 1:
                        self.exibir_mensagem("Caixa", "Valor guardado com sucesso!")
                    else:
                        self.exibir_mensagem("Caixa", f"Erro para quardar!{caixa}") 
                self.visualizar_cotas(1) 
            else:
                self.exibir_mensagem("Pagamento Cota", "Insera o valor por favor")
                self.formularioPagamento() 
    def actualizarDadoGrupo(self):
        dialog1 = formulario_passe(self)
        result1 = dialog1.exec()
        if result1 == QDialog.DialogCode.Accepted:
            resultat = caregarDadoGrupo()
            for resposta in resultat:
                senha = resposta[2]
            if senha == dialog1.linEd_senha.text():
                dialog = formulario_dadoDoGrupo(self)
                result = dialog.exec()
                if result == QDialog.DialogCode.Accepted:
                    if dialog.linEd_NomeGrupo.text() == "" or dialog.linEd_Cota.text() == "" or dialog.linEd_passe.text() == "":
                        self.exibir_mensagem("Segurança", "Preecnha todos os dados!")
                        pass
                    else:
                        actualizar = ActualizarDadoDoGrupo(1, dialog.linEd_NomeGrupo.text(), dialog.linEd_Cota.text(), dialog.linEd_passe.text())
                        if actualizar == 1:
                            self.exibir_mensagem("Segurança", "Dados actualizados com sucesso!")
                            self.Label_valorCota.setText(str(dialog.linEd_Cota.text()))
                            self.Label_titreDoGroupe.setText(str(dialog.linEd_NomeGrupo.text()))
                        else:
                            self.exibir_mensagem("Segurança", f"Erro para quardar!{actualizar}")
                else:
                    pass
            else:
                self.exibir_mensagem("Segurança", f"Palvra passe incorecto")
        else:
            pass
        
    def formularioDogrupoo(self):
        dialog = formulario_dadoDoGrupo(self)
        result = dialog.exec()
        if result == QDialog.DialogCode.Accepted:
            nome = dialog.linEd_NomeGrupo.text()
            cota = dialog.linEd_Cota.text()
            passe = dialog.linEd_passe.text()
            if nome == "":
                nome = "Grupo Coral"
            if cota == "":
                cota = 1
            if passe == "":
                passe = "0000"
            dadoDoGrupo = ActualizarDadoDoGrupo(0, nome, cota, passe)
            if dadoDoGrupo == 1:
                self.exibir_mensagem("Caixa", "Inseridos com sucesso!")
            else:
                self.exibir_mensagem("Caixa", f"Erro para quardar!{dadoDoGrupo}")
        else:
            dadoDoGrupo = ActualizarDadoDoGrupo(0, 'Grupo Coral', 0, '0000')
            if dadoDoGrupo == 1:
                self.exibir_mensagem("Caixa", "A senha provisória é: 0000")
            else:
                self.exibir_mensagem("Caixa", f"Erro para quardar!{dadoDoGrupo}")
            self.close() 

    def formularioDoMovimento(self):
        movimento = self.sender()
        dialog = formulario_caixa(self)
        result = dialog.exec()
        if result == QDialog.DialogCode.Accepted:
            if movimento.text() == "Entrada":
                entrada = str(dialog.linEd_valor.text())
                saida = 0
                mov = "entrada"
            else:
                entrada = 0
                saida = str(dialog.linEd_valor.text())
                mov = "saida"
            mov_caixa = inserir_caixa(dialog.linEd_motivo.text(), entrada, saida, mov)
            if mov_caixa == 1:
                self.exibir_mensagem("Caixa", "Movimento feito com sucesso!")
                self.visualizarCaixa()
            elif mov_caixa == 0:
                self.exibir_mensagem("Caixa", "O Saldo não é suficiente para fazer este movimento!")
            else:
                self.exibir_mensagem("Caixa", f"Erro do movimento! {mov_caixa}")

class formulario_passe(QDialog):
    def __init__(self, parent=None):
        super(formulario_passe, self).__init__(parent)
        self.setWindowTitle("Senha")

        self.linEd_senha = QLineEdit(self)
        self.linEd_senha.setEchoMode(QLineEdit.EchoMode.Password)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)     
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Preencha a sua senha"))
        layout.addWidget(self.linEd_senha)
        layout.addWidget(self.button_box)
        
class formulario_caixa(QDialog):
    def __init__(self, parent=None):
        super(formulario_caixa, self).__init__(parent)
        self.setWindowTitle("Formulario de Movimento")

        self.linEd_valor = QLineEdit(self)
        self.linEd_valor.setValidator(QIntValidator())
        #setMaxLength(4)
        self.linEd_motivo = QLineEdit(self)
        self.linEd_motivo.setPlaceholderText('Preencha o motivo')
        self.linEd_valor.setPlaceholderText('Preencha o valor')
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)     
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Formulario de Movimento"))
        layout.addWidget(self.linEd_valor)
        layout.addWidget(self.linEd_motivo)
        layout.addWidget(self.button_box)

class formulario_dadoDoGrupo(QDialog):
    def __init__(self, parent=None):
        super(formulario_dadoDoGrupo, self).__init__(parent)
        self.setWindowTitle("Definições de utilização")
        self.linEd_NomeGrupo = QLineEdit(self)
        self.linEd_Cota = QLineEdit(self)
        self.linEd_passe = QLineEdit(self)
        self.linEd_passe.setEchoMode(QLineEdit.EchoMode.Password)
        resultat = caregarDadoGrupo()
        if resultat:
            for resposta in resultat:
                self.linEd_NomeGrupo.setText(str(resposta[1]))
                self.linEd_Cota.setText(str(resposta[3]))
                self.linEd_passe.setText(str(resposta[2]))
        else:
            self.linEd_NomeGrupo.setPlaceholderText('Preencha o nome do grupo') 
            self.linEd_Cota.setPlaceholderText('Preencha o valor da cota caso tiver')
            self.linEd_passe.setPlaceholderText('Crie a senha caso precisar')
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)     
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Dados de utilizadore"))
        layout.addWidget(QLabel("Nome"))
        layout.addWidget(self.linEd_NomeGrupo)
        layout.addWidget(QLabel("Valor"))
        layout.addWidget(self.linEd_Cota)
        layout.addWidget(QLabel("Senha"))
        layout.addWidget(self.linEd_passe)
        layout.addWidget(self.button_box)

class formulario_cotas(QDialog):
    def __init__(self, parent=None):
        super(formulario_cotas, self).__init__(parent)
        self.setWindowTitle("Pagamento quotas")

        self.linEd_valor = QLineEdit(self)
        self.linEd_valor.setValidator(QIntValidator())
        self.CombBox_nome = QComboBox(self)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)     
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Selecione o nome e preencha o valor"))
        layout.addWidget(self.linEd_valor)
        layout.addWidget(self.CombBox_nome)
        layout.addWidget(self.button_box)
    
        nomes = preencher_comboBox() # function qui procura no bdd os nomes dos membros
        # Limpando o QComboBox antes de adicionar novos itens vindo do banco de dados
        self.CombBox_nome.clear()
        for id_membro, nome in nomes:
            self.CombBox_nome.addItem(nome)
            # Associar o ID como dado extra ao item do QComboBox
            self.CombBox_nome.setItemData(self.CombBox_nome.count() - 1, id_membro, Qt.ItemDataRole.UserRole)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    application = Coro_ccda()
    application.show()  
    sys.exit(app.exec())
# if __name__ == '__main__':
#     app = MyApplication(sys.argv)
#     app.lastWindowClosed.connect(app.quit)  # Evita o encerramento imediato da aplicação
#     sys.exit(app.exec())